import os
import sqlite3
import threading
import shutil
from datetime import date, datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import ttkbootstrap as tb
from ttkbootstrap.constants import *

import matplotlib

matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.patches import FancyBboxPatch

import numpy as np
import pandas as pd
import io

# Optional tkcalendar
try:
    from tkcalendar import Calendar, DateEntry

    HAS_TKCAL = True
except ImportError:
    HAS_TKCAL = False

DB = "account.db"
APP_FONT = ("微软雅黑", 12)
FONT_NORMAL = ("微软雅黑", 11)
FONT_SMALL = ("微软雅黑", 10)
TITLE_FONT = ("微软雅黑", 22, "bold")
CARD_TITLE_FONT = ("微软雅黑", 14, "bold")
MONEYPRO_COLORS = ['#4cd964', '#ff3b30', '#007aff', '#ff9500', '#5856d6', '#34c759', '#ffcc00']
APPLE_COLORS = ['#007AFF', '#34C759', '#FF9500', '#FF3B30', '#5856D6',
                '#FF2D55', '#5AC8FA', '#FFCC00', '#AF52DE', '#FF9500']

matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei']
matplotlib.rcParams['axes.unicode_minus'] = False


# ---- Database ----
def init_db(path=DB):
    first = not os.path.exists(path)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    if first:
        cur = conn.cursor()
        cur.executescript("""
        CREATE TABLE members (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);
        CREATE TABLE categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);
        CREATE TABLE records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dt TEXT NOT NULL,
            type TEXT NOT NULL,
            amount REAL NOT NULL,
            category_id INTEGER,
            payer_id INTEGER,
            note TEXT,
            created_at TEXT
        );
        """)
        defaults = ['山姆', '外食', '衣服', '车位管理费', '物业费', '车贷',
                    '水电煤气', '教育', '演出费', '医疗', '其他']
        for it in defaults:
            cur.execute("INSERT OR IGNORE INTO categories(name) VALUES(?)", (it,))
        cur.execute("INSERT OR IGNORE INTO members(name) VALUES(?)", ('小张',))
        cur.execute("INSERT OR IGNORE INTO members(name) VALUES(?)", ('小刘',))
        conn.commit()
    return conn


# ---- Repo ----
class Repo:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = conn.cursor()

    def get_members(self):
        return self.conn.execute("SELECT id,name FROM members ORDER BY id").fetchall()

    def add_member(self, name):
        self.conn.execute("INSERT OR IGNORE INTO members(name) VALUES(?)", (name,))
        self.conn.commit()

    def get_categories(self):
        return self.conn.execute("SELECT id,name FROM categories ORDER BY name").fetchall()

    def add_category(self, name):
        self.conn.execute("INSERT OR IGNORE INTO categories(name) VALUES(?)", (name,))
        self.conn.commit()

    def add_record(self, dt, type_, amount, category_id, payer_id, note):
        self.conn.execute(
            "INSERT INTO records(dt,type,amount,category_id,payer_id,note,created_at) VALUES(?,?,?,?,?,?,?)",
            (dt, type_, amount, category_id, payer_id, note, datetime.now().isoformat()))
        self.conn.commit()

    def get_record_by_id(self, record_id):
        """根据ID获取记录"""
        query = """
        SELECT r.*, c.name as category_name, m.name as payer_name
        FROM records r
        LEFT JOIN categories c ON r.category_id=c.id
        LEFT JOIN members m ON r.payer_id=m.id
        WHERE r.id=?
        """
        self.cursor.execute(query, (record_id,))
        result = self.cursor.fetchone()
        return dict(result) if result else None

    def update_record(self, record_id, dt, type_, amount, category_id, payer_id, note):
        """更新记录"""
        self.conn.execute("""
            UPDATE records 
            SET dt=?, type=?, amount=?, category_id=?, payer_id=?, note=?
            WHERE id=?
        """, (dt, type_, amount, category_id, payer_id, note, record_id))
        self.conn.commit()

    def delete_record(self, record_id):
        """删除记录"""
        self.conn.execute("DELETE FROM records WHERE id=?", (record_id,))
        self.conn.commit()

    def query_records(self, start=None, end=None, limit=500):
        sql = """SELECT r.id,r.dt,r.type,r.amount,c.name as category,m.name as payer,r.note
               FROM records r
               LEFT JOIN categories c ON r.category_id=c.id
               LEFT JOIN members m ON r.payer_id=m.id
               WHERE 1=1"""
        params = []
        if start: sql += " AND r.dt>=?"; params.append(start)
        if end: sql += " AND r.dt<=?"; params.append(end)
        sql += " ORDER BY r.dt DESC,r.id DESC LIMIT ?";
        params.append(limit)
        return self.conn.execute(sql, params).fetchall()

    def get_records_by_date_range(self, start_date, end_date):
        """按日期范围获取记录"""
        query = """
        SELECT r.id,r.dt,r.type,r.amount,c.name as category,m.name as payer,r.note
        FROM records r
        LEFT JOIN categories c ON r.category_id=c.id
        LEFT JOIN members m ON r.payer_id=m.id
        WHERE r.dt BETWEEN ? AND ?
        ORDER BY r.dt DESC, r.id DESC
        LIMIT 1000
        """
        return self.conn.execute(query, (start_date, end_date)).fetchall()

    def category_summary(self, year, month):
        sql = """SELECT c.name as category, SUM(r.amount) as total
               FROM records r
               LEFT JOIN categories c ON r.category_id=c.id
               WHERE substr(r.dt,1,7)=? AND r.type='expense'
               GROUP BY c.name
               ORDER BY total DESC"""
        return [dict(r) for r in self.conn.execute(sql, (f"{year:04d}-{month:02d}",)).fetchall()]

    def member_expense_summary(self, year, month):
        sql = """SELECT m.name as payer, SUM(r.amount) as total
               FROM records r
               LEFT JOIN members m ON r.payer_id=m.id
               WHERE substr(r.dt,1,7)=? AND r.type='expense'
               GROUP BY m.name
               ORDER BY total DESC"""
        return [dict(r) for r in self.conn.execute(sql, (f"{year:04d}-{month:02d}",)).fetchall()]

    def month_summary(self, year, month):
        """返回每个成员每种类型收入/支出总额"""
        ym = f"{year:04d}-{month:02d}"
        cur = self.conn.cursor()
        cur.execute("""
            SELECT m.name as payer, r.type, SUM(r.amount) as total
            FROM records r
            LEFT JOIN members m ON r.payer_id=m.id
            WHERE substr(r.dt,1,7)=?
            GROUP BY m.name, r.type
        """, (ym,))
        return cur.fetchall()

    def monthly_amounts_last_12(self):
        """获取最近6个月的支出数据"""
        today = date.today()
        months = []
        for i in range(11, -1, -1):  # 最近6个月
            d = (today.replace(day=1) - timedelta(days=i * 30))
            months.append({
                'year_month': d.strftime("%Y-%m"),
                'display': d.strftime("%Y年%m月"),
                'short': f"{d.month}月"
            })

        res = []
        for m in months:
            total = self.conn.execute(
                "SELECT SUM(amount) as total FROM records WHERE substr(dt,1,7)=? AND type='expense'",
                (m['year_month'],)).fetchone()['total'] or 0
            res.append({
                'month': m['display'],
                'short_month': m['short'],
                'total': total
            })
        return res

    def backup_database(self, backup_path):
        """备份数据库"""
        try:
            shutil.copy2(DB, backup_path)
            return True, "备份成功"
        except Exception as e:
            return False, f"备份失败: {str(e)}"


# ---- Settings Page ----
class SettingPage(tb.Frame):
    def __init__(self, master, db: Repo, refresh_callback=None, **kwargs):
        super().__init__(master, **kwargs)
        self.db = db
        self.refresh_callback = refresh_callback
        self._build()

    def _build(self):
        tb.Label(self, text='设置', font=TITLE_FONT).pack(anchor='w', pady=(8, 10), padx=12)

        # 成员管理
        sec = tb.Frame(self, padding=8)
        sec.pack(fill='x', padx=12, pady=6)
        tb.Label(sec, text='家庭成员管理', font=CARD_TITLE_FONT).pack(anchor='w')
        self.mem_frame = tb.Frame(sec)
        self.mem_frame.pack(fill='x', pady=6)
        self.reload_members()

        add_fr = tb.Frame(sec)
        add_fr.pack(fill='x', pady=6)
        self.new_mem_var = tk.StringVar()
        tb.Entry(add_fr, textvariable=self.new_mem_var, bootstyle="info", width=20).pack(
            side='left', expand=True, padx=(0, 8))
        tb.Button(add_fr, text='添加成员', bootstyle="primary", command=self.add_member).pack(side='left')

        # 分类管理
        sec2 = tb.Frame(self, padding=8)
        sec2.pack(fill='x', padx=12, pady=6)
        tb.Label(sec2, text='支出/收入分类管理', font=CARD_TITLE_FONT).pack(anchor='w')
        self.cat_frame = tb.Frame(sec2)
        self.cat_frame.pack(fill='x', pady=6)
        self.reload_categories()

        addc_fr = tb.Frame(sec2)
        addc_fr.pack(fill='x', pady=6)
        self.new_cat_var = tk.StringVar()
        tb.Entry(addc_fr, textvariable=self.new_cat_var, bootstyle="info", width=20).pack(
            side='left', expand=True, padx=(0, 8))
        tb.Button(addc_fr, text='添加分类', bootstyle="primary", command=self.add_category).pack(side='left')

        # 数据库备份
        sec3 = tb.Frame(self, padding=8)
        sec3.pack(fill='x', padx=12, pady=6)
        tb.Label(sec3, text='数据管理', font=CARD_TITLE_FONT).pack(anchor='w')

        backup_fr = tb.Frame(sec3)
        backup_fr.pack(fill='x', pady=10)
        tb.Button(backup_fr, text='备份数据库', bootstyle="warning",
                  command=self.backup_database, width=15).pack(side='left', padx=(0, 10))
        tb.Button(backup_fr, text='恢复数据库', bootstyle="danger",
                  command=self.restore_database, width=15).pack(side='left')

    def reload_members(self):
        for w in self.mem_frame.winfo_children():
            w.destroy()
        for m in self.db.get_members():
            fr = tb.Frame(self.mem_frame, padding=6, bootstyle="light")
            fr.pack(fill='x', pady=2)
            tb.Label(fr, text=m['name'], font=APP_FONT).pack(side='left', padx=6)

    def add_member(self):
        n = self.new_mem_var.get().strip()
        if not n:
            messagebox.showwarning("提示", "请输入成员名称")
            return
        self.db.add_member(n)
        self.new_mem_var.set('')
        self.reload_members()
        if self.refresh_callback:
            self.refresh_callback()

    def reload_categories(self):
        for w in self.cat_frame.winfo_children():
            w.destroy()
        for c in self.db.get_categories():
            fr = tb.Frame(self.cat_frame, padding=6, bootstyle="light")
            fr.pack(fill='x', pady=2)
            tb.Label(fr, text=c['name'], font=APP_FONT).pack(side='left', padx=6)

    def add_category(self):
        n = self.new_cat_var.get().strip()
        if not n:
            messagebox.showwarning("提示", "请输入分类名称")
            return
        self.db.add_category(n)
        self.new_cat_var.set('')
        self.reload_categories()
        if self.refresh_callback:
            self.refresh_callback()

    def backup_database(self):
        """备份数据库"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")],
            initialfile=f"account_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        )
        if filename:
            success, msg = self.db.backup_database(filename)
            if success:
                messagebox.showinfo("备份成功", f"数据库已备份到:\n{filename}")
            else:
                messagebox.showerror("备份失败", msg)

    def restore_database(self):
        """恢复数据库"""
        if messagebox.askyesno("确认", "恢复数据库将覆盖当前数据，是否继续？"):
            filename = filedialog.askopenfilename(
                filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")]
            )
            if filename:
                try:
                    shutil.copy2(filename, DB)
                    messagebox.showinfo("恢复成功", "数据库已恢复，请重新启动应用程序")
                except Exception as e:
                    messagebox.showerror("恢复失败", f"恢复失败: {str(e)}")


# ---- Main App ----
class App(tb.Window):
    def __init__(self):
        super().__init__(title="家庭记账", size=(1100, 720), themename="minty")
        self.conn = init_db()
        self.repo = Repo(self.conn)
        self._build()

    def get_member_emoji(self, member_name):
        """根据成员名称返回对应的emoji"""
        emoji_map = {
            '小刘': '👧',  # 女孩
            '小张': '🧑',  # 男孩
            '小张（男孩）': '🧑',
            '小张（女）': '👧',
            '小刘（女）': '👧',
            '小刘（男孩）': '🧑',
            '家庭': '🏠',
            '全家': '👨‍👩‍👧‍👦',
            '共同': '🤝',
            '合计': '💰',
            '其他': '👥',
            '公司': '🏢',
            '个人': '👤',
            '默认': '👤'
        }

        # 清理空格
        member_name_clean = member_name.strip()

        # 尝试精确匹配
        if member_name_clean in emoji_map:
            return emoji_map[member_name_clean]

        # 模糊匹配
        for key, emoji in emoji_map.items():
            if key in member_name_clean:
                return emoji

        # 根据常见关键词匹配
        if any(word in member_name_clean for word in ['女', '妈', '姐', '妹', '妻', '婆', '奶']):
            return '👧'
        elif any(word in member_name_clean for word in ['男', '爸', '哥', '弟', '夫', '公', '爷']):
            return '🧑'
        elif any(word in member_name_clean for word in ['家庭', '全家', '共同', '合计']):
            return '🏠'

        # 默认emoji
        return emoji_map['默认']

    def _build(self):
        sidebar = tb.Frame(self, width=220, bootstyle="secondary", padding=12)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tb.Label(sidebar, text="衔泥筑巢", font=TITLE_FONT, bootstyle="info").pack(pady=(6, 12))

        btns = [
            ("总览", self.show_home),
            ("记一笔", self.show_record),
            ("账单", self.show_records),
            ("统计", self.show_analysis),
            ("设置", self.show_setting)
        ]

        for t, c in btns:
            tb.Button(sidebar, text=t, width=18, bootstyle="light", command=c).pack(pady=6)

        self.body = tb.Frame(self, bootstyle="light")
        self.body.pack(side="right", fill="both", expand=True)
        self.show_home()

    def clear(self):
        for w in self.body.winfo_children():
            w.destroy()

    # --- 首页 ---
    # ---- 首页 ---
    def show_home(self):
        self.clear()

        # 统一背景：浅灰（更柔和）
        self.body.configure(style="light.TFrame")

        main_container = tb.Frame(self.body, style="light.TFrame")
        main_container.pack(fill='both', expand=True)

        # 顶部标题区域 —— 改为柔和渐变蓝
        top_frame = tb.Frame(main_container, bootstyle="info", padding=(20, 15))
        top_frame.pack(fill='x')

        tb.Label(top_frame, text="🏠 家庭总览",
                 font=("微软雅黑", 26, "bold"),
                 bootstyle="inverse-info").pack(anchor='center', pady=(0, 5))

        today = date.today()
        tb.Label(top_frame, text=f"📅 {today.year}年{today.month}月{today.day}日",
                 font=("微软雅黑", 12),
                 bootstyle="inverse-info").pack(anchor='center')

        # 内容区域背景换成淡白
        content_frame = tb.Frame(main_container, padding=(20, 15), bootstyle="light")
        content_frame.pack(fill='both', expand=True)

        rows = self.repo.month_summary(today.year, today.month)
        df = pd.DataFrame(rows, columns=['payer', 'type', 'total']) if rows else pd.DataFrame(
            columns=['payer', 'type', 'total'])

        members = [r['name'] for r in self.repo.get_members()]

        # 顶部统计卡片
        stats_frame = tb.Frame(content_frame, style="light.TFrame")
        stats_frame.pack(fill='x', pady=(0, 25))

        month_income = sum(r['total'] for r in rows if r['type'] == 'income')
        month_expense = sum(r['total'] for r in rows if r['type'] == 'expense')
        month_balance = month_income - month_expense

        # --- 总收入卡片：柔和绿色 ---
        income_card = tb.Frame(stats_frame, padding=20, bootstyle="success")
        income_card.pack(side='left', fill='both', expand=True, padx=(0, 12))

        income_header = tb.Frame(income_card, bootstyle="success")
        income_header.pack(fill='x', pady=(0, 12))

        tb.Label(income_header, text="💰", font=("微软雅黑", 22),
                 bootstyle="inverse-success").pack(side='left', padx=(0, 10))
        tb.Label(income_header, text="总收入", font=("微软雅黑", 14, "bold"),
                 bootstyle="inverse-success").pack(side='left')

        tb.Label(income_card, text=f"¥{month_income:,.2f}",
                 font=("微软雅黑", 20, "bold"),
                 bootstyle="inverse-success").pack(anchor='center', pady=(0, 5))

        tb.Label(income_card, text="本月累计收入", font=("微软雅黑", 10),
                 bootstyle="inverse-success").pack(anchor='center')

        # --- 总支出卡片：柔和红色 ---
        expense_card = tb.Frame(stats_frame, padding=20, bootstyle="danger")
        expense_card.pack(side='left', fill='both', expand=True, padx=(0, 12))

        expense_header = tb.Frame(expense_card, bootstyle="danger")
        expense_header.pack(fill='x', pady=(0, 12))

        tb.Label(expense_header, text="💸", font=("微软雅黑", 22),
                 bootstyle="inverse-danger").pack(side='left', padx=(0, 10))
        tb.Label(expense_header, text="总支出", font=("微软雅黑", 14, "bold"),
                 bootstyle="inverse-danger").pack(side='left')

        tb.Label(expense_card, text=f"¥{month_expense:,.2f}",
                 font=("微软雅黑", 20, "bold"),
                 bootstyle="inverse-danger").pack(anchor='center', pady=(0, 5))

        tb.Label(expense_card, text="本月累计支出", font=("微软雅黑", 10),
                 bootstyle="inverse-danger").pack(anchor='center')

        # --- 结余卡片：蓝色正向、黄色警告 ---
        if month_balance >= 0:
            balance_style = "info"
            balance_emoji = "📈"
            balance_title = "结余"
        else:
            balance_style = "warning"
            balance_emoji = "⚠️"
            balance_title = "超支"

        balance_card = tb.Frame(stats_frame, padding=20, bootstyle=balance_style)
        balance_card.pack(side='left', fill='both', expand=True)

        balance_header = tb.Frame(balance_card, bootstyle=balance_style)
        balance_header.pack(fill='x', pady=(0, 12))

        tb.Label(balance_header, text=balance_emoji, font=("微软雅黑", 22),
                 bootstyle=f"inverse-{balance_style}").pack(side='left', padx=(0, 10))
        tb.Label(balance_header, text=f"本月{balance_title}",
                 font=("微软雅黑", 14, "bold"),
                 bootstyle=f"inverse-{balance_style}").pack(side='left')

        tb.Label(balance_card, text=f"¥{month_balance:,.2f}",
                 font=("微软雅黑", 20, "bold"),
                 bootstyle=f"inverse-{balance_style}").pack(anchor='center', pady=(0, 5))

        tb.Label(balance_card, text="收入 - 支出",
                 font=("微软雅黑", 10),
                 bootstyle=f"inverse-{balance_style}").pack(anchor='center')

        # ---- 成员统计 ----
        tb.Label(content_frame, text="👨‍👩‍👧‍👦 家庭成员统计",
                 font=("微软雅黑", 16, "bold"),
                 bootstyle="info").pack(anchor='w', pady=(20, 12))

        members_container = tb.Frame(content_frame)
        members_container.pack(fill='both', expand=True)

        emoji_map = {
            '小刘': '👩',
            '小张': '🧑',
            '家庭': '🏠',
            '共同': '🤝',
            '其他': '👤'
        }

        # 成员卡片配色（更柔和、更优雅）
        member_styles = ["secondary", "info", "warning", "success", "danger"]

        # --- 网格布局成员卡片 ---
        if len(members) > 2:
            grid_container = tb.Frame(members_container)
            grid_container.pack(fill='both', expand=True)

            for i, member in enumerate(members):
                style = member_styles[i % len(member_styles)]
                card = tb.Frame(grid_container, padding=15, bootstyle=style)

                row = i // 2
                col = i % 2
                card.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')

                grid_container.columnconfigure(col, weight=1)
                grid_container.rowconfigure(row, weight=1)

                self._create_member_card_content(card, member, df, emoji_map)

        else:
            # 横排
            row_container = tb.Frame(members_container)
            row_container.pack(fill='x', expand=True)

            for i, member in enumerate(members):
                style = member_styles[i % len(member_styles)]
                card = tb.Frame(row_container, padding=15, bootstyle=style)
                card.pack(side='left', fill='both', expand=True, padx=10)

                self._create_member_card_content(card, member, df, emoji_map)

        # 底部横条
        bottom_frame = tb.Frame(main_container, bootstyle="light", padding=(10, 10))
        bottom_frame.pack(fill='x', side='bottom')

        tb.Label(bottom_frame, text="💡 家庭理财，幸福生活",
                 font=("微软雅黑", 11),
                 bootstyle="secondary").pack(pady=5)

    def _create_member_card_content(self, card, member, df, emoji_map):
        """创建成员卡片内容（辅助函数）"""
        # 获取成员emoji
        member_emoji = emoji_map.get(member, '👤')

        # 成员名称区域
        name_frame = tb.Frame(card)
        name_frame.pack(fill='x', pady=(0, 15))

        tb.Label(name_frame, text=member_emoji,
                 font=("微软雅黑", 24)).pack(side='left', padx=(0, 10))

        tb.Label(name_frame, text=member,
                 font=("微软雅黑", 14, "bold"),
                 bootstyle="primary").pack(side='left')

        # 计算收入和支出
        income = 0
        expense = 0

        if not df.empty:
            member_df = df[df['payer'] == member]
            income = member_df.loc[member_df['type'] == 'income', 'total'].sum()
            expense = member_df.loc[member_df['type'] == 'expense', 'total'].sum()

        member_balance = income - expense

        # 收入显示
        income_frame = tb.Frame(card)
        income_frame.pack(fill='x', pady=8)

        tb.Label(income_frame, text="收入:",
                 font=APP_FONT,
                 bootstyle="secondary").pack(side='left')

        tb.Label(income_frame, text=f"¥{income:,.2f}",
                 font=("微软雅黑", 14, "bold"),
                 bootstyle="success").pack(side='right')

        # 支出显示
        expense_frame = tb.Frame(card)
        expense_frame.pack(fill='x', pady=8)

        tb.Label(expense_frame, text="支出:",
                 font=APP_FONT,
                 bootstyle="secondary").pack(side='left')

        tb.Label(expense_frame, text=f"¥{expense:,.2f}",
                 font=("微软雅黑", 14, "bold"),
                 bootstyle="danger").pack(side='right')

        # 分隔线
        separator = ttk.Separator(card, orient='horizontal')
        separator.pack(fill='x', pady=15)

        # 结余显示
        balance_frame = tb.Frame(card)
        balance_frame.pack(fill='x')

        tb.Label(balance_frame, text="结余:",
                 font=("微软雅黑", 12, "bold"),
                 bootstyle="secondary").pack(side='left')

        # 根据结余选择颜色
        if member_balance >= 0:
            balance_style = "success"
        else:
            balance_style = "warning"

        tb.Label(balance_frame, text=f"¥{member_balance:,.2f}",
                 font=("微软雅黑", 16, "bold"),
                 bootstyle=balance_style).pack(side='right')

    # ---------- 记一笔 ----------
    def show_record(self):
        self.clear()
        main = tb.Frame(self.body, padding=12)
        main.pack(fill="both", expand=True)

        left = tb.Frame(main)
        left.pack(side='left', fill='y', expand=True, padx=(0, 20))

        # 收支模式
        self.mode = tk.StringVar(value='expense')
        sw = tb.Frame(left)
        sw.pack(pady=8)

        self.icon_label = tk.Label(sw, text="🛒", font=("微软雅黑", 16))
        self.icon_label.grid(row=0, column=0, padx=6)

        def set_mode(mode):
            self.mode.set(mode)
            self.icon_label.config(text="💰" if mode == "income" else "🛒")
            self.update_categories()

        tb.Button(sw, text="支出", bootstyle="danger-outline", width=12,
                  command=lambda: set_mode('expense')).grid(row=0, column=1, padx=6)
        tb.Button(sw, text="收入", bootstyle="success-outline", width=12,
                  command=lambda: set_mode('income')).grid(row=0, column=2, padx=6)

        # 分类
        tb.Label(left, text="分类", font=APP_FONT).pack(anchor='w', pady=(12, 4))
        self.cat_cb = ttk.Combobox(left, width=22, font=("微软雅黑", 11))
        self.cat_cb.pack()
        self.income_categories = ['工资', '奖金', '理财收益', '兼职收入', '其他收入']
        self.update_categories()

        # 金额
        tb.Label(left, text="金额", font=APP_FONT).pack(anchor='w', pady=(12, 4))
        self.amount_entry = tb.Entry(left, font=("微软雅黑", 14), justify='center', width=18)
        self.amount_entry.pack(pady=6)

        # 付款人
        tb.Label(left, text="付款人", font=APP_FONT).pack(anchor='w', pady=(12, 4))
        mems = [r['name'] for r in self.repo.get_members()]
        self.payer_cb = ttk.Combobox(left, values=mems, width=20, font=("微软雅黑", 11))
        self.payer_cb.pack()
        if mems:
            self.payer_cb.current(0)

        # 日期 - 简化布局，直接输入 + 提示
        tb.Label(left, text="日期", font=APP_FONT).pack(anchor='w', pady=(12, 4))
        date_fr = tb.Frame(left)
        date_fr.pack(fill='x')

        inner_fr = tb.Frame(date_fr)
        inner_fr.pack(anchor='center')  # 居中

        self.dt_var = tk.StringVar(value=date.today().isoformat())
        self.dt_entry = tb.Entry(inner_fr, textvariable=self.dt_var,
                                 font=("微软雅黑", 11), width=15)
        self.dt_entry.pack(side='left', padx=(0, 5))

        def show_date_hint():
            """显示日期格式提示"""
            messagebox.showinfo("日期格式", "请输入日期，格式为：YYYY-MM-DD\n例如：2025-12-01")

        tb.Button(inner_fr, text="📅", bootstyle="secondary",
                  width=3, command=show_date_hint).pack(side='left')

        # 备注
        tb.Label(left, text="备注", font=APP_FONT).pack(anchor='w', pady=(12, 4))
        self.note_txt = tk.Text(left, height=2, font=("微软雅黑", 11))
        self.note_txt.pack(fill='x')

        # 按钮
        btn_frame = tb.Frame(left)
        btn_frame.pack(side='bottom', fill='x', pady=(20, 12))
        action_frame = tb.Frame(btn_frame)
        action_frame.pack(anchor='center')

        tb.Button(action_frame, text="保存", bootstyle="primary", width=16,
                  command=self.on_save).pack(side='left', padx=(0, 10))
        tb.Button(action_frame, text="清空", bootstyle="secondary", width=16,
                  command=lambda: self.clear_record_form()).pack(side='left')

        # 最近记录
        right = tb.Frame(main)
        right.pack(side='right', fill='both', expand=True)
        tb.Label(right, text="最近记录", font=CARD_TITLE_FONT).pack(anchor='w')

        self.recent_list = tk.Listbox(right, height=16, font=("微软雅黑", 11))
        self.recent_list.pack(fill='both', expand=True, pady=6)
        self.reload_recent()

    def clear_record_form(self):
        """清空记录表单"""
        self.amount_entry.delete(0, 'end')
        self.note_txt.delete('1.0', 'end')
        # 重置分类和付款人为默认值
        if hasattr(self, 'cat_cb') and self.cat_cb['values']:
            try:
                self.cat_cb.current(0)
            except Exception:
                pass
        if hasattr(self, 'payer_cb') and self.payer_cb['values']:
            try:
                self.payer_cb.current(0)
            except Exception:
                pass

    def update_categories(self):
        # when mode changes or categories updated
        all_categories = [r['name'] for r in self.repo.get_categories()]
        if self.mode.get() == 'income':
            values = self.income_categories
        else:
            values = all_categories

        self.cat_cb['values'] = values
        if values:
            try:
                self.cat_cb.current(0)
            except Exception:
                pass

    def reload_recent(self):
        self.recent_list.delete(0, 'end')
        rows = self.repo.query_records(limit=20)
        for r in rows:
            text = f"{r['dt']} {r['type']} {r['amount']:.2f} {r['category'] or ''} {r['payer'] or ''}"
            self.recent_list.insert('end', text)

    def on_save(self):
        amt = self.amount_entry.get().strip()
        try:
            amtf = float(amt)
            if amtf <= 0:
                messagebox.showerror("错误", "金额必须大于0")
                return
        except Exception:
            messagebox.showerror("错误", "请输入合法金额")
            return

        dt = self.dt_var.get().strip()
        # basic date validation
        try:
            _ = datetime.fromisoformat(dt)
        except Exception:
            messagebox.showerror("错误", "日期格式请使用 YYYY-MM-DD")
            return

        typ = self.mode.get()
        cat_name = self.cat_cb.get().strip()
        payer_name = self.payer_cb.get().strip()
        note = self.note_txt.get("1.0", "end").strip()

        if not cat_name or not payer_name:
            messagebox.showerror("错误", "请选择分类和付款人")
            return

        cur = self.conn.cursor()
        cur.execute("SELECT id FROM categories WHERE name=?", (cat_name,))
        c = cur.fetchone()
        cid = c['id'] if c else None

        if not cid:
            # 如果分类不存在，创建新分类
            cur.execute("INSERT OR IGNORE INTO categories(name) VALUES(?)", (cat_name,))
            cid = cur.lastrowid
            self.conn.commit()

        cur.execute("SELECT id FROM members WHERE name=?", (payer_name,))
        m = cur.fetchone()
        mid = m['id'] if m else None

        self.repo.add_record(dt, typ, amtf, cid, mid, note)
        messagebox.showinfo("成功", "已保存")
        self.amount_entry.delete(0, 'end')
        self.note_txt.delete('1.0', 'end')
        self.reload_recent()

    # ---------- 账单列表 ----------
    def show_records(self):
        self.clear()

        # 标题
        tb.Label(self.body, text="账单列表", font=TITLE_FONT, bootstyle="info").pack(
            anchor='w', pady=12, padx=12)

        # 筛选区域
        filter_frame = tb.Frame(self.body, padding=10)
        filter_frame.pack(fill='x', padx=12)

        tb.Label(filter_frame, text="起始日期", font=APP_FONT).grid(row=0, column=0, padx=(0, 5))

        # 默认显示最近90天
        start_date = (date.today() - timedelta(days=90)).isoformat()
        end_date = date.today().isoformat()

        start_var = tk.StringVar(value=start_date)
        end_var = tk.StringVar(value=end_date)

        tb.Entry(filter_frame, textvariable=start_var, width=12).grid(row=0, column=1, padx=(0, 10))
        tb.Label(filter_frame, text="至", font=APP_FONT).grid(row=0, column=2, padx=5)
        tb.Entry(filter_frame, textvariable=end_var, width=12).grid(row=0, column=3, padx=(0, 10))

        # 创建一个容器用于显示记录
        container = tb.Frame(self.body)
        container.pack(fill='both', expand=True, padx=12, pady=(0, 12))

        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(container, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient='vertical', command=canvas.yview)

        # 创建滚动框架
        scrollable_frame = tb.Frame(canvas)

        # 配置Canvas
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.bind('<Configure>', on_canvas_configure)
        scrollable_frame.bind('<Configure>', on_frame_configure)

        # 放置Canvas和Scrollbar
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        def apply_filter():
            # 清除现有记录
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            s = start_var.get().strip()
            e = end_var.get().strip()

            if not s or not e:
                messagebox.showerror("错误", "请输入起始日期和结束日期")
                return

            rows = self.repo.get_records_by_date_range(s, e)

            if not rows:
                tb.Label(scrollable_frame, text="暂无记录", font=APP_FONT, bootstyle="secondary").pack(pady=20)
                return

            # 显示记录
            for r in rows:
                self._create_record_item(scrollable_frame, r)

        tb.Button(filter_frame, text="筛选", bootstyle="primary", command=apply_filter).grid(
            row=0, column=4, padx=(10, 0))

        # 初始加载
        apply_filter()

    def _create_record_item(self, parent, record):
        """创建单个记录显示项"""
        frame = tb.Frame(parent, padding=10)
        frame.pack(fill='x', pady=2, padx=2)
        frame.configure(bootstyle="light")

        # 左半部分：信息
        info_frame = tb.Frame(frame)
        info_frame.pack(side='left', fill='both', expand=True)

        # 日期和类型
        type_text = "收入" if record['type'] == 'income' else "支出"
        type_color = "success" if record['type'] == 'income' else "danger"

        tb.Label(info_frame, text=f"{record['dt']}  {type_text}",
                 font=("微软雅黑", 12, "bold"), bootstyle=type_color).pack(anchor='w')

        # 金额、分类、付款人
        tb.Label(info_frame,
                 text=f"金额: ¥{record['amount']:.2f}    分类: {record['category'] or ''}    付款人: {record['payer'] or ''}",
                 font=APP_FONT).pack(anchor='w', pady=(4, 0))

        # 备注（如果有）
        if record['note']:
            tb.Label(info_frame, text=f"备注: {record['note']}",
                     font=("微软雅黑", 10), bootstyle="secondary").pack(anchor='w', pady=(2, 0))

        # 右半部分：按钮
        btn_frame = tb.Frame(frame)
        btn_frame.pack(side='right', fill='y')

        tb.Button(btn_frame, text="编辑", bootstyle="warning", width=6,
                  command=lambda rid=record['id']: self.edit_record_dialog(rid)).pack(pady=2)

        tb.Button(btn_frame, text="删除", bootstyle="danger", width=6,
                  command=lambda rid=record['id']: self.delete_record(rid)).pack(pady=2)

    def delete_record(self, record_id):
        """删除记录"""
        if messagebox.askyesno("确认删除", "确定要删除这条记录吗？"):
            self.repo.delete_record(record_id)
            messagebox.showinfo("成功", "记录已删除")
            self.show_records()

    # ---------- 编辑记录对话框 ----------
    # ---------- 编辑记录对话框 ----------
    # ---------- 编辑记录对话框 ----------
    def edit_record_dialog(self, record_id):
        """编辑记录"""
        record = self.repo.get_record_by_id(record_id)
        if not record:
            messagebox.showerror("错误", "记录不存在")
            return

        # 创建对话框 - 使用更合适的尺寸
        dialog = tb.Toplevel(self)
        dialog.title("编辑账单记录")
        dialog.geometry("480x650")  # 调整尺寸，更紧凑
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        # 对话框居中显示
        dialog.update_idletasks()
        width = 480
        height = 650
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')

        # 主容器
        main_frame = tb.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)

        # 标题区域
        tb.Label(main_frame, text="编辑账单记录",
                 font=("微软雅黑", 16, "bold"), bootstyle="info").pack(pady=(0, 20))

        # 表单容器 - 使用简单的布局
        form_frame = tb.Frame(main_frame)
        form_frame.pack(fill='both', expand=True)

        # 类型选择 - 水平布局
        type_frame = tb.Frame(form_frame)
        type_frame.pack(fill='x', pady=(0, 15))

        tb.Label(type_frame, text="类型:",
                 font=("微软雅黑", 12), width=6).pack(side='left')

        type_var = tk.StringVar(value=record['type'])
        type_btn_frame = tb.Frame(type_frame)
        type_btn_frame.pack(side='left', fill='x', expand=True)

        tb.Radiobutton(type_btn_frame, text="支出", variable=type_var,
                       value="expense", bootstyle="danger").pack(side='left', padx=(0, 15))
        tb.Radiobutton(type_btn_frame, text="收入", variable=type_var,
                       value="income", bootstyle="success").pack(side='left')

        # 金额 - 简洁布局
        amount_frame = tb.Frame(form_frame)
        amount_frame.pack(fill='x', pady=(0, 15))

        tb.Label(amount_frame, text="金额:",
                 font=("微软雅黑", 12), width=6).pack(side='left')

        amount_var = tk.StringVar(value=str(record['amount']))
        amount_entry = tb.Entry(amount_frame, textvariable=amount_var,
                                font=("微软雅黑", 12), width=15)
        amount_entry.pack(side='left', padx=(0, 5))
        tb.Label(amount_frame, text="元", font=("微软雅黑", 12)).pack(side='left')

        # 分类 - 紧凑布局
        category_frame = tb.Frame(form_frame)
        category_frame.pack(fill='x', pady=(0, 15))

        tb.Label(category_frame, text="分类:",
                 font=("微软雅黑", 12), width=6).pack(side='left')

        # 获取分类
        all_categories = [row['name'] for row in self.repo.get_categories()]
        income_categories = ['工资', '奖金', '理财收益', '兼职收入', '其他收入']

        current_categories = income_categories if record['type'] == 'income' else all_categories
        category_var = tk.StringVar()

        category_combo = ttk.Combobox(category_frame, textvariable=category_var,
                                      values=current_categories, font=("微软雅黑", 12),
                                      width=15, state="readonly")  # 缩短宽度
        category_combo.pack(side='left')

        # 设置当前分类
        current_category = record.get('category_name', '')
        if current_category in current_categories:
            category_var.set(current_category)
        elif current_categories:
            category_var.set(current_categories[0])

        # 成员 - 紧凑布局
        member_frame = tb.Frame(form_frame)
        member_frame.pack(fill='x', pady=(0, 15))

        tb.Label(member_frame, text="成员:",
                 font=("微软雅黑", 12), width=6).pack(side='left')

        members = [row['name'] for row in self.repo.get_members()]
        member_var = tk.StringVar()

        member_combo = ttk.Combobox(member_frame, textvariable=member_var,
                                    values=members, font=("微软雅黑", 12),
                                    width=15, state="readonly")  # 缩短宽度
        member_combo.pack(side='left')

        # 设置当前成员
        current_member = record.get('payer_name', '')
        if current_member:
            member_var.set(current_member)
        elif members:
            member_var.set(members[0])

        # 日期 - 简化布局，避免复杂的日历选择器
        date_frame = tb.Frame(form_frame)
        date_frame.pack(fill='x', pady=(0, 15))

        tb.Label(date_frame, text="日期:",
                 font=("微软雅黑", 12), width=6).pack(side='left')

        date_var = tk.StringVar(value=record['dt'])
        date_entry = tb.Entry(date_frame, textvariable=date_var,
                              font=("微软雅黑", 12), width=15)
        date_entry.pack(side='left', padx=(0, 5))

        # 简单的日期选择提示
        def show_date_hint():
            """显示日期格式提示"""
            messagebox.showinfo("日期格式", "请输入日期，格式为：YYYY-MM-DD\n例如：2025-12-01")

        tb.Button(date_frame, text="📅", bootstyle="secondary",
                  command=show_date_hint, width=3).pack(side='left')

        # 备注 - 简化布局
        note_frame = tb.Frame(form_frame)
        note_frame.pack(fill='both', expand=True, pady=(0, 15))

        tb.Label(note_frame, text="备注:",
                 font=("微软雅黑", 12)).pack(anchor='w', pady=(0, 5))

        # 备注文本框 - 更短
        note_text_frame = tb.Frame(note_frame)
        note_text_frame.pack(fill='both', expand=True)

        note_text = tk.Text(note_text_frame, height=2, font=("微软雅黑", 12), wrap='word')
        note_scrollbar = ttk.Scrollbar(note_text_frame, orient='vertical', command=note_text.yview)

        note_text.pack(side='left', fill='both', expand=True)
        note_scrollbar.pack(side='right', fill='y')

        note_text.config(yscrollcommand=note_scrollbar.set)
        note_text.insert('1.0', record['note'] or '')

        # 定义更新分类函数
        def update_categories(*args):
            if type_var.get() == 'income':
                new_categories = income_categories
            else:
                new_categories = all_categories

            category_combo['values'] = new_categories
            if new_categories:
                category_var.set(new_categories[0])

        type_var.trace('w', update_categories)

        # 添加分隔线
        separator = ttk.Separator(form_frame, orient='horizontal')
        separator.pack(fill='x', pady=(15, 20))

        # 按钮区域 - 居中显示
        button_frame = tb.Frame(form_frame)
        button_frame.pack(fill='x')

        def save():
            try:
                amount = float(amount_var.get().strip())
                if amount <= 0:
                    messagebox.showerror("错误", "金额必须大于0")
                    return
            except ValueError:
                messagebox.showerror("错误", "请输入有效的金额")
                return

            dt = date_var.get().strip()
            typ = type_var.get()
            cat_name = category_var.get().strip()
            payer_name = member_var.get().strip()
            note = note_text.get('1.0', 'end-1c').strip()

            # 验证
            if not dt or not cat_name or not payer_name:
                messagebox.showerror("错误", "请填写完整信息")
                return

            try:
                datetime.strptime(dt, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("错误", "日期格式应为 YYYY-MM-DD")
                return

            # 获取分类ID
            cur = self.conn.cursor()
            cur.execute('SELECT id FROM categories WHERE name = ?', (cat_name,))
            category_row = cur.fetchone()

            if not category_row:
                # 如果分类不存在，创建新分类
                cur.execute('INSERT INTO categories (name) VALUES (?)', (cat_name,))
                self.conn.commit()
                category_id = cur.lastrowid
            else:
                category_id = category_row['id']

            # 获取成员ID
            cur.execute('SELECT id FROM members WHERE name = ?', (payer_name,))
            member_row = cur.fetchone()
            if not member_row:
                messagebox.showerror("错误", "成员不存在")
                return
            payer_id = member_row['id']

            # 更新记录
            try:
                self.repo.update_record(record_id, dt, typ, amount, category_id, payer_id, note)
                messagebox.showinfo("成功", "记录已更新")
                dialog.destroy()
                self.show_records()
            except Exception as e:
                messagebox.showerror("错误", f"更新失败: {str(e)}")

        # 按钮容器 - 居中显示
        btn_container = tb.Frame(button_frame)
        btn_container.pack(expand=True)

        tb.Button(btn_container, text="保存", bootstyle="primary",
                  command=save, width=10).pack(side='left', padx=(0, 10))

        tb.Button(btn_container, text="取消", bootstyle="secondary",
                  command=dialog.destroy, width=10).pack(side='left')

        # 确保对话框显示时自动聚焦到金额输入框
        def on_dialog_shown():
            amount_entry.focus_set()
            amount_entry.select_range(0, 'end')

        dialog.after(100, on_dialog_shown)
    # ---------- 统计分析 ----------
    # ---------- 统计分析 ----------
    def show_analysis(self):
        self.clear()
        self.current_view = "analysis"

        # 创建主容器
        main_container = tb.Frame(self.body)
        main_container.pack(fill='both', expand=True)

        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(main_container, bg="#f8f8f8")
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tb.Frame(canvas)  # 这里定义scrollable_frame

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", on_canvas_configure)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 标题
        tb.Label(scrollable_frame, text="统计分析", font=("微软雅黑", 16, "bold"), bootstyle="info").pack(
            anchor='nw', pady=8, padx=12)

        today = date.today()

        # 月份选择
        month_frame = tb.Frame(scrollable_frame)
        month_frame.pack(fill='x', padx=12, pady=(0, 15))

        tb.Label(month_frame, text="选择月份：", font=APP_FONT).pack(side='left')

        self.stats_year = tk.IntVar(value=today.year)
        self.stats_month = tk.IntVar(value=today.month)

        tb.Spinbox(month_frame, from_=2020, to=2030, textvariable=self.stats_year, width=8).pack(side='left',
                                                                                                 padx=(0, 5))
        tb.Label(month_frame, text="年", font=APP_FONT).pack(side='left')
        tb.Spinbox(month_frame, from_=1, to=12, textvariable=self.stats_month, width=6).pack(side='left', padx=(5, 5))
        tb.Label(month_frame, text="月", font=APP_FONT).pack(side='left', padx=(0, 10))

        tb.Button(month_frame, text="刷新", bootstyle="primary",
                  command=self.refresh_analysis, width=10).pack(side='left')

        # 获取月度统计数据
        year = self.stats_year.get()
        month = self.stats_month.get()

        # 月度汇总数据
        monthly_data = self.repo.month_summary(year, month)
        df = pd.DataFrame(monthly_data, columns=['payer', 'type', 'total']) if monthly_data else pd.DataFrame(
            columns=['payer', 'type', 'total'])

        # 计算总计
        total_income = df.loc[df['type'] == 'income', 'total'].sum() if not df.empty else 0
        total_expense = df.loc[df['type'] == 'expense', 'total'].sum() if not df.empty else 0
        balance = total_income - total_expense

        # 获取最大支出分类
        cat_summary = self.repo.category_summary(year, month)
        max_category = cat_summary[0]['category'] if cat_summary else "无"
        max_category_amount = cat_summary[0]['total'] if cat_summary else 0

        # 获取最小支出成员
        member_summary = self.repo.member_expense_summary(year, month)
        min_member = member_summary[-1]['payer'] if member_summary else "无"
        min_member_amount = member_summary[-1]['total'] if member_summary else 0

        # 统计卡片
        stats_cards_frame = tb.Frame(scrollable_frame)
        stats_cards_frame.pack(fill='x', padx=12, pady=(0, 20))

        # 第一行卡片
        row1 = tb.Frame(stats_cards_frame)
        row1.pack(fill='x', pady=(0, 10))

        # 总收入卡片
        income_card = tb.Frame(row1, padding=15, bootstyle="light", relief="solid", width=200)
        income_card.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 图标 + 标题
        income_title_frame = tb.Frame(income_card)
        income_title_frame.pack(anchor='w', fill='x')
        tb.Label(income_title_frame, text="💰", font=("微软雅黑", 16)).pack(side='left', padx=(0, 8))
        tb.Label(income_title_frame, text="总收入", font=APP_FONT).pack(side='left')

        tb.Label(income_card, text=f"¥{total_income:.2f}",
                 font=("微软雅黑", 18, "bold"), bootstyle="success").pack(anchor='w', pady=(5, 0))

        # 总支出卡片
        expense_card = tb.Frame(row1, padding=15, bootstyle="light", relief="solid", width=200)
        expense_card.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 图标 + 标题
        expense_title_frame = tb.Frame(expense_card)
        expense_title_frame.pack(anchor='w', fill='x')
        tb.Label(expense_title_frame, text="💸", font=("微软雅黑", 16)).pack(side='left', padx=(0, 8))
        tb.Label(expense_title_frame, text="总支出", font=APP_FONT).pack(side='left')

        tb.Label(expense_card, text=f"¥{total_expense:.2f}",
                 font=("微软雅黑", 18, "bold"), bootstyle="danger").pack(anchor='w', pady=(5, 0))

        # 月度结余卡片
        balance_card = tb.Frame(row1, padding=15, bootstyle="light", relief="solid", width=200)
        balance_card.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 图标 + 标题
        balance_title_frame = tb.Frame(balance_card)
        balance_title_frame.pack(anchor='w', fill='x')
        tb.Label(balance_title_frame, text="💹", font=("微软雅黑", 16)).pack(side='left', padx=(0, 8))
        tb.Label(balance_title_frame, text="月度结余", font=APP_FONT).pack(side='left')

        balance_color = "success" if balance >= 0 else "danger"
        tb.Label(balance_card, text=f"¥{balance:.2f}",
                 font=("微软雅黑", 18, "bold"), bootstyle=balance_color).pack(anchor='w', pady=(5, 0))

        # 第二行卡片
        row2 = tb.Frame(stats_cards_frame)
        row2.pack(fill='x')

        # 最大支出分类卡片
        max_cat_card = tb.Frame(row2, padding=15, bootstyle="light", relief="solid", width=200)
        max_cat_card.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 图标 + 标题
        max_cat_title_frame = tb.Frame(max_cat_card)
        max_cat_title_frame.pack(anchor='w', fill='x')
        tb.Label(max_cat_title_frame, text="📊", font=("微软雅黑", 16)).pack(side='left', padx=(0, 8))
        tb.Label(max_cat_title_frame, text="最大支出分类", font=APP_FONT).pack(side='left')

        tb.Label(max_cat_card, text=max_category, font=APP_FONT).pack(anchor='w', pady=(5, 0))
        tb.Label(max_cat_card, text=f"¥{max_category_amount:.2f}",
                 font=("微软雅黑", 14, "bold"), bootstyle="danger").pack(anchor='w', pady=(5, 0))

        # 最小支出成员卡片
        min_mem_card = tb.Frame(row2, padding=15, bootstyle="light", relief="solid", width=200)
        min_mem_card.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 图标 + 标题
        min_mem_title_frame = tb.Frame(min_mem_card)
        min_mem_title_frame.pack(anchor='w', fill='x')

        # 根据成员名称选择不同的图标
        if min_member == '小张':
            member_icon = "👤"  # 男孩
        elif min_member == '小刘':
            member_icon = "👩"  # 女孩
        elif '公用' in min_member or '共同' in min_member or '家庭' in min_member:
            member_icon = "🏠"  # 家庭
        else:
            member_icon = "👤"  # 默认

        tb.Label(min_mem_title_frame, text=member_icon, font=("微软雅黑", 16)).pack(side='left', padx=(0, 8))
        tb.Label(min_mem_title_frame, text="最小支出成员", font=APP_FONT).pack(side='left')

        tb.Label(min_mem_card, text=min_member, font=APP_FONT).pack(anchor='w', pady=(5, 0))
        tb.Label(min_mem_card, text=f"¥{min_member_amount:.2f}",
                 font=("微软雅黑", 14, "bold"), bootstyle="warning").pack(anchor='w', pady=(5, 0))

        # ---- 支出类别统计 ----
        cat_frame = tb.Frame(scrollable_frame, padding=15, bootstyle="light", relief="solid")
        cat_frame.pack(fill='x', padx=12, pady=(0, 15))

        tb.Label(cat_frame, text="📋 支出类别明细",
                 font=("微软雅黑", 14, "bold"), bootstyle="info").pack(anchor='w', pady=(0, 15))

        # 获取支出类别数据
        month_total = total_expense

        # 将category_summary转换为DataFrame
        if cat_summary:
            df_cat = pd.DataFrame(cat_summary)
        else:
            df_cat = pd.DataFrame()

        if not df_cat.empty and month_total > 0:
            # 创建表格框架
            table_frame = tb.Frame(cat_frame)
            table_frame.pack(fill='x')

            # 添加表头
            header_frame = tb.Frame(table_frame)
            header_frame.pack(fill='x', pady=(0, 5))

            tb.Label(header_frame, text="类别", font=("微软雅黑", 11, "bold"),
                     width=15, anchor='w').pack(side='left', padx=(0, 10))
            tb.Label(header_frame, text="金额", font=("微软雅黑", 11, "bold"),
                     width=15, anchor='center').pack(side='left', padx=(0, 10))
            tb.Label(header_frame, text="占比", font=("微软雅黑", 11, "bold"),
                     width=10, anchor='center').pack(side='left')

            # 添加分隔线
            separator = ttk.Separator(table_frame, orient='horizontal')
            separator.pack(fill='x', pady=5)

            # 添加数据行
            data_frame = tb.Frame(table_frame)
            data_frame.pack(fill='x')

            # 按金额排序
            df_cat_sorted = df_cat.sort_values('total', ascending=False)

            # 定义颜色
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4',
                      '#FFEAA7', '#DDA0DD', '#98D8C8', '#F7DC6F']

            for idx, row in df_cat_sorted.iterrows():
                amount = row['total']
                percentage = (amount / month_total * 100) if month_total > 0 else 0
                color = colors[idx % len(colors)] if idx < len(colors) else '#888888'

                row_frame = tb.Frame(data_frame, padding=(10, 8))
                row_frame.pack(fill='x', pady=2)

                # 颜色标记（使用圆点）
                color_label = tb.Label(row_frame, text="●", font=("Arial", 18),
                                       foreground=color, width=2)
                color_label.pack(side='left')

                # 类别名称
                tb.Label(row_frame, text=row['category'], font=("微软雅黑", 11),
                         width=20, anchor='w').pack(side='left', padx=(5, 0))

                # 金额
                tb.Label(row_frame, text=f"¥{amount:,.2f}", font=("微软雅黑", 11),
                         width=15, anchor='center').pack(side='left')

                # 百分比和进度条
                perc_frame = tb.Frame(row_frame)
                perc_frame.pack(side='left', fill='x', expand=True)

                tb.Label(perc_frame, text=f"{percentage:.1f}%",
                         font=("微软雅黑", 10), width=6).pack(side='left')

                # 进度条
                progress_frame = tb.Frame(perc_frame, height=8)
                progress_frame.pack(side='left', fill='x', expand=True, padx=(5, 0))
                progress_frame.pack_propagate(False)

                # 创建进度条（使用Frame模拟）
                progress_width = min(100, percentage * 2)  # 控制最大宽度
                progress = tb.Frame(progress_frame, height=8, bootstyle="info")
                progress.place(x=0, y=0, width=progress_width, relheight=1.0)

                # 背景
                bg = tb.Frame(progress_frame, height=8, bootstyle="secondary")
                bg.place(x=0, y=0, relwidth=1.0, relheight=1.0)

                # 将进度条放在背景前面
                progress.lift()
        else:
            tb.Label(cat_frame, text="本月暂无支出数据",
                     font=("微软雅黑", 12), bootstyle="secondary").pack(pady=20)

        # ---- 成员支出统计 ----
        member_frame = tb.Frame(scrollable_frame, padding=15, bootstyle="light", relief="solid")
        member_frame.pack(fill='x', padx=12, pady=(0, 15))

        tb.Label(member_frame, text="👥 成员支出明细",
                 font=("微软雅黑", 14, "bold"), bootstyle="primary").pack(anchor='w', pady=(0, 15))

        # 将member_summary转换为DataFrame
        if member_summary:
            df_mem = pd.DataFrame(member_summary)
        else:
            df_mem = pd.DataFrame()

        if not df_mem.empty:
            # 计算成员支出总额
            member_total = df_mem['total'].sum()

            # 创建表格框架
            table_frame2 = tb.Frame(member_frame)
            table_frame2.pack(fill='x')

            # 添加表头
            header_frame2 = tb.Frame(table_frame2)
            header_frame2.pack(fill='x', pady=(0, 5))

            tb.Label(header_frame2, text="成员", font=("微软雅黑", 11, "bold"),
                     width=15, anchor='w').pack(side='left', padx=(0, 10))
            tb.Label(header_frame2, text="金额", font=("微软雅黑", 11, "bold"),
                     width=15, anchor='center').pack(side='left', padx=(0, 10))
            tb.Label(header_frame2, text="占比", font=("微软雅黑", 11, "bold"),
                     width=10, anchor='center').pack(side='left')

            # 添加分隔线
            separator2 = ttk.Separator(table_frame2, orient='horizontal')
            separator2.pack(fill='x', pady=5)

            # 添加数据行
            data_frame2 = tb.Frame(table_frame2)
            data_frame2.pack(fill='x')

            # 按金额排序
            df_mem_sorted = df_mem.sort_values('total', ascending=False)

            # 成员颜色
            member_colors = ['#007AFF', '#34C759', '#FF9500']

            for idx, row in df_mem_sorted.iterrows():
                amount = row['total']
                percentage = (amount / member_total * 100) if member_total > 0 else 0

                # 根据成员选择颜色
                if row['payer'] == '家庭':
                    color = member_colors[0]
                    icon = "🏠"
                elif row['payer'] == '小刘':
                    color = member_colors[1]
                    icon = "👩"
                elif row['payer'] == '小张':
                    color = member_colors[2]
                    icon = "👤"
                else:
                    color = '#888888'
                    icon = "👤"

                row_frame = tb.Frame(data_frame2, padding=(10, 8))
                row_frame.pack(fill='x', pady=2)

                # 头像/图标
                icon_label = tb.Label(row_frame, text=icon, font=("Arial", 12),
                                      width=2)
                icon_label.pack(side='left')

                # 成员名称
                tb.Label(row_frame, text=row['payer'], font=("微软雅黑", 11),
                         width=20, anchor='w').pack(side='left', padx=(5, 0))

                # 金额
                tb.Label(row_frame, text=f"¥{amount:,.2f}", font=("微软雅黑", 11),
                         width=15, anchor='center').pack(side='left')

                # 百分比和进度条
                perc_frame = tb.Frame(row_frame)
                perc_frame.pack(side='left', fill='x', expand=True)

                tb.Label(perc_frame, text=f"{percentage:.1f}%",
                         font=("微软雅黑", 10), width=6).pack(side='left')

                # 进度条
                progress_frame = tb.Frame(perc_frame, height=8)
                progress_frame.pack(side='left', fill='x', expand=True, padx=(5, 0))
                progress_frame.pack_propagate(False)

                # 创建进度条
                progress_width = min(100, percentage * 2)

                # 根据成员选择进度条样式
                if row['payer'] == '家庭':
                    progress_style = "primary"
                elif row['payer'] == '小刘':
                    progress_style = "warning"
                elif row['payer'] == '小张':
                    progress_style = "info"
                else:
                    progress_style = "success"

                progress = tb.Frame(progress_frame, height=8, bootstyle=progress_style)
                progress.place(x=0, y=0, width=progress_width, relheight=1.0)

                # 背景
                bg = tb.Frame(progress_frame, height=8, bootstyle="secondary")
                bg.place(x=0, y=0, relwidth=1.0, relheight=1.0)

                progress.lift()
        else:
            tb.Label(member_frame, text="本月暂无成员支出数据",
                     font=("微软雅黑", 12), bootstyle="secondary").pack(pady=20)

        # ---- 月度趋势（12个月，重新设计布局） ----
        # ---- 月度支出趋势（居中布局，优化统计摘要） ----
        trend_frame = tb.Frame(scrollable_frame, padding=15, bootstyle="light", relief="solid")
        trend_frame.pack(fill='x', padx=12, pady=(0, 15))

        tb.Label(trend_frame, text="📈 月度支出趋势（最近12个月）",
                 font=("微软雅黑", 14, "bold"), bootstyle="warning").pack(anchor='w', pady=(0, 15))

        # 获取最近12个月的数据
        monthly_trend_data = []

        for i in range(11, -1, -1):
            d = (today.replace(day=1) - timedelta(days=i * 30))
            year_month = d.strftime("%Y-%m")

            query = "SELECT SUM(amount) as total FROM records WHERE substr(dt,1,7)=? AND type='expense'"
            result = self.conn.execute(query, (year_month,)).fetchone()
            total = result['total'] if result and result['total'] else 0

            monthly_trend_data.append({
                'date': d,
                'month': d.strftime("%Y年%m月"),
                'short_month': f"{d.month}月",
                'year_month': f"{d.year}-{d.month:02d}",
                'total': total
            })

        if monthly_trend_data and any(item['total'] > 0 for item in monthly_trend_data):
            # 计算统计数据
            non_zero_data = [d for d in monthly_trend_data if d['total'] > 0]

            if non_zero_data:
                avg_expense = sum(d['total'] for d in non_zero_data) / len(non_zero_data)
                max_data = max(non_zero_data, key=lambda x: x['total'])
                min_data = min(non_zero_data, key=lambda x: x['total'])
            else:
                avg_expense = 0
                max_data = {'short_month': '无', 'total': 0}
                min_data = {'short_month': '无', 'total': 0}

            # 创建主容器
            main_container = tb.Frame(trend_frame)
            main_container.pack(fill='both', expand=True)

            # 左侧图表区域
            chart_area = tb.Frame(main_container)
            chart_area.pack(side='left', fill='both', expand=True, padx=(0, 10))

            # 图表容器 - 固定高度
            chart_container = tb.Frame(chart_area, height=380)
            chart_container.pack(fill='both', expand=True)
            chart_container.pack_propagate(False)

            # 在容器中创建Canvas
            chart_canvas = tk.Canvas(chart_container, bg='white', highlightthickness=0)
            chart_canvas.pack(fill='both', expand=True)

            # 获取Canvas的实际尺寸
            def draw_chart():
                canvas_width = chart_canvas.winfo_width()
                canvas_height = chart_canvas.winfo_height()

                if canvas_width <= 1 or canvas_height <= 1:
                    # 如果Canvas还未绘制，稍后重试
                    chart_container.after(100, draw_chart)
                    return

                # 清空Canvas
                chart_canvas.delete("all")

                # 边距设置
                left_margin = 60  # Y轴标签
                right_margin = 30
                top_margin = 50
                bottom_margin = 70  # X轴标签

                # 绘图区域
                plot_left = left_margin
                plot_top = top_margin
                plot_bottom = canvas_height - bottom_margin
                plot_width = canvas_width - left_margin - right_margin
                plot_height = plot_bottom - plot_top

                # 计算最大值
                max_value = max(item['total'] for item in monthly_trend_data) if monthly_trend_data else 1
                if max_value == 0:
                    max_value = 1

                # 1. 绘制Y轴和刻度
                # Y轴线
                chart_canvas.create_line(plot_left, plot_top, plot_left, plot_bottom, width=2, fill='#333333')

                # 绘制Y轴刻度线（5个刻度）
                y_steps = 5
                for i in range(y_steps + 1):
                    y = plot_bottom - (i * plot_height / y_steps)
                    value = (i * max_value / y_steps)

                    # 刻度线
                    chart_canvas.create_line(plot_left - 5, y, plot_left, y, width=1, fill='#666666')

                    # 网格线
                    chart_canvas.create_line(plot_left, y, plot_left + plot_width, y, fill='#f0f0f0', width=1)

                    # Y轴标签 - 格式化金额
                    if value >= 1000000:
                        label = f"¥{value / 1000000:.1f}M"
                    elif value >= 100000:
                        label = f"¥{value / 100000:.1f}M"
                    elif value >= 10000:
                        label = f"¥{value / 10000:.1f}万"
                    elif value >= 1000:
                        label = f"¥{value / 1000:.0f}k"
                    else:
                        label = f"¥{value:.0f}"

                    chart_canvas.create_text(plot_left - 10, y, text=label,
                                             font=("微软雅黑", 9), fill='#666666', anchor='e')

                # 2. 绘制X轴
                chart_canvas.create_line(plot_left, plot_bottom, plot_left + plot_width, plot_bottom,
                                         width=2, fill='#333333')

                # 3. 绘制柱状图 - 居中显示
                bar_width = 30
                spacing = 15
                total_bars = 12
                total_width_needed = (total_bars * bar_width) + ((total_bars - 1) * spacing)

                # 如果总宽度超过绘图区域，调整柱宽和间距
                if total_width_needed > plot_width:
                    scale_factor = plot_width / total_width_needed
                    bar_width = int(bar_width * scale_factor)
                    spacing = int(spacing * scale_factor)
                    total_width_needed = (total_bars * bar_width) + ((total_bars - 1) * spacing)

                # 计算起始位置，使图表居中
                x_offset = plot_left + (plot_width - total_width_needed) // 2

                # 定义颜色
                colors = ['#4A90E2', '#5AA469', '#F5A623', '#D0021B',
                          '#9013FE', '#50E3C2', '#F8E71C', '#8B572A',
                          '#417505', '#BD10E0', '#7ED321', '#4A4A4A']

                # 绘制每个月的柱状图
                for i, data in enumerate(monthly_trend_data):
                    x = x_offset + i * (bar_width + spacing)

                    # 计算柱高
                    height = (data['total'] / max_value) * plot_height
                    y_top = plot_bottom - height

                    # 柱状图颜色
                    color = colors[i % len(colors)]

                    # 绘制柱状图
                    if data['total'] > 0:
                        chart_canvas.create_rectangle(x, y_top, x + bar_width, plot_bottom,
                                                      fill=color, outline='white', width=1)

                        # 添加立体感 - 使用白色线条
                        chart_canvas.create_line(x, y_top, x + bar_width, y_top,
                                                 fill='#FFFFFF', width=1)
                        chart_canvas.create_line(x, y_top, x, plot_bottom,
                                                 fill='#FFFFFF', width=1)
                    else:
                        # 无数据时显示灰色虚线
                        chart_canvas.create_rectangle(x, plot_bottom - 2, x + bar_width, plot_bottom,
                                                      fill='#f0f0f0', outline='#cccccc', width=1, dash=(2, 2))

                    # X轴标签 - 月份
                    month_text = data['short_month']

                    # 绘制月份标签
                    label_y = plot_bottom + 15
                    chart_canvas.create_text(x + bar_width / 2, label_y,
                                             text=month_text,
                                             font=("微软雅黑", 10),
                                             fill='#333333', anchor='n')

                # 4. 当前月份高亮边框
                current_month_index = next((i for i, data in enumerate(monthly_trend_data)
                                            if data['date'].year == today.year and data['date'].month == today.month),
                                           -1)

                if current_month_index >= 0:
                    current_data = monthly_trend_data[current_month_index]
                    x_current = x_offset + current_month_index * (bar_width + spacing)

                    # 添加橙色边框标识当前月份
                    chart_canvas.create_rectangle(x_current - 3, plot_top - 5,
                                                  x_current + bar_width + 3, plot_bottom + 5,
                                                  outline='#FF9500', width=2, dash=(3, 3))

            # 在Canvas绘制完成后调用draw_chart
            chart_canvas.bind('<Configure>', lambda e: draw_chart())

            # 右侧统计摘要区域
            stats_area = tb.Frame(main_container, width=200)
            stats_area.pack(side='right', fill='y')
            stats_area.pack_propagate(False)

            # 统计摘要标题
            tb.Label(stats_area, text="统计摘要",
                     font=("微软雅黑", 12, "bold"), bootstyle="info").pack(anchor='w', pady=(0, 15))

            # 格式化金额函数
            def format_amount(amount):
                if amount >= 1000000:
                    return f"¥{amount / 1000000:.1f}M"
                elif amount >= 100000:
                    return f"¥{amount / 100000:.1f}M"
                elif amount >= 10000:
                    return f"¥{amount / 10000:.1f}万"
                elif amount >= 1000:
                    return f"¥{amount / 1000:.0f}k"
                else:
                    return f"¥{amount:.0f}"

            # 平均支出
            avg_frame = tb.Frame(stats_area)
            avg_frame.pack(fill='x', pady=(0, 20))
            tb.Label(avg_frame, text="平均支出：",
                     font=("微软雅黑", 11), bootstyle="secondary", anchor='w').pack(anchor='w')
            tb.Label(avg_frame, text=format_amount(avg_expense),
                     font=("微软雅黑", 14, "bold"), bootstyle="info").pack(anchor='w', pady=(5, 0))

            # 最高支出
            max_frame = tb.Frame(stats_area)
            max_frame.pack(fill='x', pady=(0, 20))
            tb.Label(max_frame, text="最高支出：",
                     font=("微软雅黑", 11), bootstyle="secondary", anchor='w').pack(anchor='w')

            max_value_text = f"{max_data['short_month']} {format_amount(max_data['total'])}"
            tb.Label(max_frame, text=max_value_text,
                     font=("微软雅黑", 14, "bold"), bootstyle="danger").pack(anchor='w', pady=(5, 0))

            # 最低支出
            min_frame = tb.Frame(stats_area)
            min_frame.pack(fill='x', pady=(0, 20))
            tb.Label(min_frame, text="最低支出：",
                     font=("微软雅黑", 11), bootstyle="secondary", anchor='w').pack(anchor='w')

            min_value_text = f"{min_data['short_month']} {format_amount(min_data['total'])}"
            tb.Label(min_frame, text=min_value_text,
                     font=("微软雅黑", 14, "bold"), bootstyle="success").pack(anchor='w', pady=(5, 0))

            # 图例说明 - 放在统计摘要下方
            legend_frame = tb.Frame(stats_area)
            legend_frame.pack(fill='x', pady=(20, 0))

            tb.Label(legend_frame, text="图例说明：",
                     font=("微软雅黑", 11), bootstyle="secondary").pack(anchor='w', pady=(0, 10))

            # 创建垂直排列的图例
            legend_items = [
                ("● 柱状图表", "表示月度支出总额"),
                ("● 虚线", "为年份分隔线"),
                ("● 橙色边框", "标识当前月份")
            ]

            for symbol, description in legend_items:
                item_frame = tb.Frame(legend_frame)
                item_frame.pack(fill='x', pady=(0, 5))

                tb.Label(item_frame, text=symbol,
                         font=("微软雅黑", 9), bootstyle="secondary", width=10, anchor='w').pack(side='left')
                tb.Label(item_frame, text=description,
                         font=("微软雅黑", 9), bootstyle="secondary", anchor='w').pack(side='left', padx=(5, 0))

        else:
            # 无数据时的显示
            no_data_frame = tb.Frame(trend_frame, padding=30)
            no_data_frame.pack(fill='both', expand=True)

            tb.Label(no_data_frame, text="📊",
                     font=("微软雅黑", 48), bootstyle="secondary").pack(pady=(0, 15))

            tb.Label(no_data_frame, text="暂无月度趋势数据",
                     font=("微软雅黑", 12, "bold"), bootstyle="secondary").pack(pady=(0, 10))

            tb.Label(no_data_frame, text="记录更多支出数据后，这里将显示12个月的支出趋势图表",
                     font=("微软雅黑", 10), bootstyle="secondary", wraplength=400).pack()

        # ---- 导出功能 ----
        export_frame = tb.Frame(scrollable_frame)
        export_frame.pack(fill='x', padx=15, pady=(0, 20))

        def export_simple_excel():
            try:
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel 文件", "*.xlsx")],
                    initialfile=f"家庭账单统计_{year}年{month}月.xlsx"
                )

                if filename:
                    # 创建DataFrame
                    export_data = []

                    # 添加统计摘要
                    export_data.append({
                        '类型': '统计摘要',
                        '项目': '总收入',
                        '金额': total_income,
                        '月份': f"{year}年{month}月"
                    })

                    export_data.append({
                        '类型': '统计摘要',
                        '项目': '总支出',
                        '金额': total_expense,
                        '月份': f"{year}年{month}月"
                    })

                    export_data.append({
                        '类型': '统计摘要',
                        '项目': '月度结余',
                        '金额': balance,
                        '月份': f"{year}年{month}月"
                    })

                    export_data.append({
                        '类型': '统计摘要',
                        '项目': '最大支出分类',
                        '金额': max_category_amount,
                        '备注': max_category,
                        '月份': f"{year}年{month}月"
                    })

                    # 添加支出类别
                    if not df_cat.empty:
                        for _, row in df_cat.iterrows():
                            export_data.append({
                                '类型': '支出类别',
                                '项目': row['category'],
                                '金额': row['total'],
                                '月份': f"{year}年{month}月"
                            })

                    # 添加成员支出
                    if not df_mem.empty:
                        for _, row in df_mem.iterrows():
                            export_data.append({
                                '类型': '成员支出',
                                '项目': row['payer'],
                                '金额': row['total'],
                                '月份': f"{year}年{month}月"
                            })

                    # 添加月度趋势
                    for data in monthly_trend_data:
                        if data['total'] > 0:
                            export_data.append({
                                '类型': '月度趋势',
                                '项目': data['month'],
                                '金额': data['total'],
                                '月份': data['month']
                            })

                    if export_data:
                        df_export = pd.DataFrame(export_data)

                        # 使用pandas的ExcelWriter，设置中文引擎
                        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                            df_export.to_excel(writer, index=False, sheet_name=f'{year}年{month}月统计')

                            # 调整列宽
                            worksheet = writer.sheets[f'{year}年{month}月统计']
                            worksheet.column_dimensions['A'].width = 15
                            worksheet.column_dimensions['B'].width = 20
                            worksheet.column_dimensions['C'].width = 15
                            worksheet.column_dimensions['D'].width = 20

                        messagebox.showinfo("导出成功", f"数据已保存到:\n{filename}")
                    else:
                        messagebox.showwarning("无数据", "没有数据可以导出")

            except Exception as e:
                messagebox.showerror("导出失败", f"保存文件时出错:\n{str(e)}")

        tb.Button(export_frame, text="📥 导出Excel", bootstyle="success",
                  command=export_simple_excel, width=20).pack(pady=10)

    def refresh_analysis(self):
        """刷新统计分析"""
        self.show_analysis()

    def refresh_analysis(self):
        """刷新统计分析"""
        self.show_analysis()

    def export_excel(self):
        """导出Excel"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = openpyxl.Workbook()

            # 获取当前月份数据
            year = self.stats_year.get()
            month = self.stats_month.get()

            # 成员收入工作表
            ws1 = wb.active
            ws1.title = "成员收入"

            income_rows = self.repo.month_summary(year, month)

            if income_rows:
                income_data = []
                for row in income_rows:
                    income_data.append({
                        'payer': row['payer'],
                        'type': row['type'],
                        'total': row['total']
                    })

                df_income = pd.DataFrame(income_data)
                df_income = df_income[df_income['type'] == 'income']

                if not df_income.empty:
                    for r in dataframe_to_rows(df_income, index=False, header=True):
                        ws1.append(r)
                else:
                    ws1.append(["暂无收入数据"])
            else:
                ws1.append(["暂无收入数据"])

            # 支出明细工作表
            ws2 = wb.create_sheet("支出明细")

            rows = self.repo.query_records(limit=1000)

            if rows:
                expense_data = []
                for row in rows:
                    if row['type'] == 'expense':
                        expense_data.append({
                            'date': row['dt'],
                            'amount': row['amount'],
                            'category': row['category'] or '',
                            'payer': row['payer'] or '',
                            'note': row['note'] or ''
                        })

                df_exp = pd.DataFrame(expense_data)

                if not df_exp.empty:
                    headers = ['日期', '金额', '分类', '付款人', '备注']
                    ws2.append(headers)

                    for _, row in df_exp.iterrows():
                        ws2.append([
                            row['date'],
                            row['amount'],
                            row['category'],
                            row['payer'],
                            row['note']
                        ])
                else:
                    ws2.append(["暂无支出数据"])
            else:
                ws2.append(["暂无支出数据"])

            # 统计汇总工作表
            ws3 = wb.create_sheet("统计汇总")

            cat_rows = self.repo.category_summary(year, month)
            df_cat = pd.DataFrame(cat_rows) if cat_rows else pd.DataFrame()

            ws3.append(["支出类别", "金额", "占比"])

            if not df_cat.empty:
                total_expense = df_cat['total'].sum()
                for i, row in df_cat.iterrows():
                    percentage = (row['total'] / total_expense * 100) if total_expense > 0 else 0
                    ws3.append([row['category'], row['total'], f"{percentage:.1f}%"])
            else:
                ws3.append(["暂无统计数据"])

            # 成员支出占比工作表
            ws4 = wb.create_sheet("成员支出占比")

            member_rows = self.repo.member_expense_summary(year, month)
            df_mem = pd.DataFrame(member_rows) if member_rows else pd.DataFrame()

            ws4.append(["成员", "支出金额", "占比"])

            if not df_mem.empty:
                total_member_expense = df_mem['total'].sum()
                for i, row in df_mem.iterrows():
                    percentage = (row['total'] / total_member_expense * 100) if total_member_expense > 0 else 0
                    ws4.append([row['payer'], row['total'], f"{percentage:.1f}%"])
            else:
                ws4.append(["暂无成员支出数据"])

            # 保存文件
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile=f"账单统计_{date.today().strftime('%Y%m%d')}.xlsx"
            )

            if filename:
                wb.save(filename)
                messagebox.showinfo("导出成功", f"Excel 已保存到:\n{filename}")

        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件时出错:\n{str(e)}")


    # ---------- 设置 ----------
    def show_setting(self):
        self.clear()
        page = SettingPage(self.body, self.repo, refresh_callback=self._on_settings_changed)
        page.pack(fill='both', expand=True)

    def _on_settings_changed(self):
        """设置更改后的回调函数"""
        try:
            if hasattr(self, 'cat_cb'):
                self.update_categories_in_record()
        except Exception:
            pass

    def update_categories_in_record(self):
        """更新记录页面的分类和成员"""
        if hasattr(self, 'cat_cb'):
            all_categories = [r['name'] for r in self.repo.get_categories()]
            if hasattr(self, 'mode') and self.mode.get() == 'income':
                values = self.income_categories
            else:
                values = all_categories
            self.cat_cb['values'] = values
            if values:
                try:
                    self.cat_cb.current(0)
                except Exception:
                    pass
        if hasattr(self, 'payer_cb'):
            mems = [r['name'] for r in self.repo.get_members()]
            self.payer_cb['values'] = mems
            if mems:
                try:
                    self.payer_cb.current(0)
                except Exception:
                    pass


if __name__ == "__main__":
    app = App()
    app.mainloop()